import pywinauto
from pywinauto import Application
from pywinauto.keyboard import send_keys
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import re
import openpyxl


return_workbook = Workbook()
return_sheet = return_workbook.active
Input_Array = []
user_data = []

# This program only works if the users computer is actively running Label Maintenance off Epicor
# Make sure Label Maintenance is the front most window open as well
class Automator:


    def __init__(self):
        self.invalid_index = 2
        self.existed_index = 2

        # region UI Prompts

        # Get input file
        input_file = self.get_user_input("Enter input excel file path (in file format and must have .xlsx at the "
                                         "end): ", self.is_valid_file)
        user_data.append(input_file)

        # Get sheet index
        sheet_name = self.get_user_input_sheet("Enter sheet name: ", self.sheet_exists, input_file)

        sheet_index = self.get_sheet_index(input_file, sheet_name)
        user_data.append(sheet_index)

        # Get column letter
        column_letter = self.get_user_input("Enter column letter: ", self.is_valid_column)
        user_data.append(column_letter)

        # Get first row
        first_row = self.get_user_input("Enter first row: ", self.is_valid_integer)
        user_data.append(int(first_row))

        # Get last row
        last_row = self.get_user_input("Enter last row: ", self.is_valid_integer)
        user_data.append(int(last_row))


        # endregion UI Prompts

        self.file_name = Automator.create_excel(self)


    def run_test1(self):
        try:
            # Connect the application to the running Label Maintenance
            app = Application(backend="uia").connect(title="Label Maintenance")
            print('Connection to Label Maintenance achieved')

            # Load in a specific workbook
            book = load_workbook(user_data[0])

            # Load in user data
            last_row = user_data[4]
            first_row = user_data[3]
            col_letter = user_data[2]
            n = user_data[1]
            sheets = book.sheetnames
            ws = book[sheets[n]]


            # Loop through the given parts and read/write the part number into Label Maintenance
            # The first number in the range function is the starting row
            # The second number is the final row
            for i in range(first_row, last_row + 1):
                app = Application(backend="uia").connect(title="Label Maintenance")
                main_window = app.window(title='Label Maintenance')

                # Find the cell and get its value
                cell = col_letter + str(i)
                cell_value = ws[cell].value

                # Type cell value into text box
                main_window.child_window(auto_id='txtKeyField').type_keys(cell_value)
                send_keys("{TAB}")

                # Check for invalid part numbers
                if main_window.child_window(title='Error').exists(timeout=10):
                    main_window.child_window(title='OK').click()
                    if main_window.child_window(title='Error').exists(timeout=5):
                        main_window.child_window(title='OK').click()
                    # Write invalid part into Excel file
                    Automator.write_invalid(self, cell_value, self.file_name)
                    print(cell_value + " - Invalid Part Number")
                    continue

                # Confirms label does not exist
                if main_window.child_window(title="Add New Confirmation").exists(timeout=5):

                    main_window.child_window(auto_id='btnYes2').click_input()

                    # Check for invalid part numbers
                    if main_window.child_window(title='Error').exists(timeout=10):
                        main_window.child_window(title='OK').click()
                        if main_window.child_window(title='Error').exists(timeout=5):
                            main_window.child_window(title='OK').click()

                        # Write invalid part number in Excel sheet
                        Automator.write_invalid(self, cell_value, self.file_name)
                        print(str(cell_value) + " - Invalid Part Number")
                        continue

                    else:
                        # Save if part number is valid
                        main_window.child_window(title="Save").click_input()
                        main_window.child_window(title="Clear").click_input()
                        print(str(cell_value) + " - Label Created")
                else:
                    Automator.write_exists(self, cell_value, self.file_name)
                    main_window.child_window(title="Clear").click_input()
                    print(str(cell_value) + " - Label Already Exists")

        except pywinauto.findwindows.ElementNotFoundError:
            print("Connection Failed. Please be sure to have Label Maintenance running")
            response = input("Try again? (yes or no)")
            if response == "yes":
                self.run_test1()

        except Exception as e:
            print(e)
            raise e

        print("Session summary is saved in Label Creator -> build")

    # region File Reading
    def create_file_name(self) -> str:
        # Save the workbook with a unique name (e.g., based on timestamp)
        file_name = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return file_name

    def create_excel(self) -> str:
        # Write data to the pre-created sheet (global variable)
        return_sheet['A1'] = 'Invalid PN'
        return_sheet['B1'] = 'Already Had Label'

        # Give the sheet a unique name and save it
        file_name = Automator.create_file_name(self)
        return_workbook.save(file_name)
        print(f"Excel file '{file_name}' created successfully.")

        return file_name

    def write_invalid(self, part_number, file_name):
        # Write PN into Excel file
        return_sheet['A' + str(self.invalid_index)] = part_number
        self.invalid_index += 1
        return_workbook.save(file_name)

    def write_exists(self, part_number, file_name):
        # Write PN into Excel file
        return_sheet['B' + str(self.existed_index)] = part_number
        self.existed_index += 1
        return_workbook.save(file_name)

    # endregion File Reading

    # region UI Methods
    def is_valid_file(self, file_path):
        return os.path.isfile(file_path) and file_path.lower().endswith('.xlsx')

    def is_valid_column(self, column):
        return re.match(r'^[A-Za-z]+$', column)

    def is_valid_integer(self, value):
        return value.isdigit()

    def get_user_input(self, prompt, validator_func):
        while True:
            user_input = input(prompt)
            if validator_func(user_input):
                return user_input
            else:
                print("Invalid input. Please try again.")

    def get_user_input_sheet(self, prompt, validator_func, input_file):
        while True:
            user_input = input(prompt)
            if validator_func(input_file, user_input):
                return user_input
            else:
                print("Invalid input. Please try again.")


    def sheet_exists(self, excel_file_path, sheet_name):
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            return sheet_name in workbook.sheetnames
        except FileNotFoundError:
            print(f"File not found: {excel_file_path}")
            return False
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Invalid Excel file: {excel_file_path}")
            return False
        except Exception as e:
            print("An error occured. Please try again.")


    def get_sheet_index(self, excel_file_path, sheet_name):
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
            sheet_index = workbook.sheetnames.index(sheet_name)
            return sheet_index
        except FileNotFoundError:
            print(f"File not found: {excel_file_path}")
            return False
        except openpyxl.utils.exceptions.InvalidFileException:
            print(f"Invalid Excel file: {excel_file_path}")
            return False
        except ValueError:
            print(f"Sheet '{sheet_name}' not found in the Excel file.")
            return False

        return None

    # endregion UI Methods








