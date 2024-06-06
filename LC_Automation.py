import pywinauto
from pywinauto import Application
from pywinauto.keyboard import send_keys
from openpyxl import Workbook, load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from datetime import datetime
import re
import os
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from pywinauto.timings import Timings
import sys

return_workbook = Workbook()
return_sheet = return_workbook.active
Input_Array = []
user_data = []


# This program only works if the users computer is actively running Label Maintenance off Epicor
# Make sure Label Maintenance is the front most window open as well
class Automator:

    def __init__(self):
        self.invalid_index = 2
        self.existed_index = 2

        # region UI Prompts
        root = tk.Tk()
        root.title("Label Creator - File Information")
        root.geometry("500x400")  # Window size

        # Labels
        lblInputFile = tk.Label(root, text="Input File", padx=10, pady=10)
        lblSheetIndex = tk.Label(root, text="Sheet Name", padx=10, pady=10)
        lblColumnLetter = tk.Label(root, text="Column Letter", padx=10, pady=10)
        lblFirstRow = tk.Label(root, text="First Row #", padx=10, pady=10)
        lblLastRow = tk.Label(root, text="Last Row #", padx=10, pady=10)

        # Text boxes
        txtInputFile = tk.Entry(root)
        txtSheetName = tk.Entry(root)
        txtColumnLetter = tk.Entry(root)
        txtFirstRow = tk.Entry(root)
        txtLastRow = tk.Entry(root)

        # region UI Methods

        def browse_file():
            file_path = filedialog.askopenfilename()
            txtInputFile.delete(0, tk.END)
            txtInputFile.insert(0, file_path)

        def get_textbox_contents():
            try:

                # Validate Input File
                if txtInputFile.get():
                    if validate_file_location(txtInputFile.get().strip()):
                        user_data.append(txtInputFile.get().strip())
                    else:
                        messagebox.showerror("Error", "Invalid input file")
                        return

                    # Validate Sheet
                    if sheet_exists(txtInputFile.get(), txtSheetName.get().strip()):
                        index = get_sheet_index(txtInputFile.get().strip(), txtSheetName.get().strip())
                        user_data.append(index)
                    else:
                        messagebox.showerror("Error", "Invalid sheet name")
                        return
                else:
                    messagebox.showerror("Error", "Please list an input file")
                    return

                # Validate Column Letter
                if is_valid_column(txtColumnLetter.get().strip()):
                    user_data.append(txtColumnLetter.get().strip())
                else:
                    messagebox.showerror("Error", "Invalid column letter")
                    return

                # Validate Rows
                if is_valid_integer(txtFirstRow.get().strip()):
                    if is_valid_integer(txtLastRow.get().strip()):
                        if int(txtFirstRow.get()) < int(txtLastRow.get().strip()):
                            user_data.append(txtFirstRow.get().strip())
                        else:
                            messagebox.showerror("Error", "Invalid row order")
                            return
                    else:
                        messagebox.showerror("Error", "Last row is invalid")
                        return
                else:
                    messagebox.showerror("Error", "First row is invalid")
                    return

                user_data.append(txtLastRow.get().strip())

                # Checks user_data for any empty values
                if len(user_data) != 5:
                    print("Please enter information in all text boxes.")
                else:
                    print(user_data)
                    root.destroy()
            except Exception as e:
                print(e)

        def is_valid_column(column):
            return re.match(r'^[A-Za-z]+$', column)

        def is_valid_integer(value):
            return value.isdigit()

        def sheet_exists(excel_file_path, sheet_name):
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
                return sheet_name in workbook.sheetnames
            except FileNotFoundError:
                print(f"File not found: {excel_file_path}")
                return False
            except openpyxl.utils.exceptions.InvalidFileException:
                print(f"Invalid Excel file: {excel_file_path}")
                return False
            except Exception(BaseException) as e:
                print("An error occured. Please try again.")

        def get_sheet_index(excel_file_path, sheet_name):
            try:
                workbook = openpyxl.load_workbook(excel_file_path)
                sheet_index = workbook.sheetnames.index(sheet_name)
                return sheet_index
            except FileNotFoundError:
                print(f"File not found: {excel_file_path}")
                return False
            except openpyxl.utils.exceptions.InvalidFileException:
                print(f"Invalid Excel file: {excel_file_path}")
                return False
            except ValueError:
                print(f"Sheet '{sheet_name}' not found in the Excel file.")
                return False

        def validate_file_location(file_path):
            """
            Validates if the provided file path is valid.

            Args:
                file_path (str): The file path to validate.

            Returns:
                bool: True if the file path is valid, False otherwise.
            """
            # Check if the file path is not empty
            if not file_path:
                return False

            # Check if the file path exists
            if not os.path.exists(file_path):
                return False

            # Check if the file path points to a file (not a directory)
            if os.path.isdir(file_path):
                return False

            return True

        def on_closing():
            if messagebox.askokcancel("Quit", "Do you want to quit?"):
                root.destroy()
                sys.exit()

        # endregion UI Methods

        # Bind the window close event to the custom method
        root.protocol("WM_DELETE_WINDOW", on_closing)

        # Buttons
        browse_button = tk.Button(root, text="Browse", command=browse_file, padx=7, pady=7)
        submit_button = tk.Button(root, text="Submit", command=get_textbox_contents, padx=7, pady=7)

        # Grid layout
        lblInputFile.grid(row=0, column=0, sticky="w", padx=7, pady=7)
        txtInputFile.grid(row=0, column=1, padx=7, pady=7)
        lblSheetIndex.grid(row=1, column=0, sticky="w", padx=7, pady=7)
        txtSheetName.grid(row=1, column=1, padx=7, pady=7)
        lblColumnLetter.grid(row=2, column=0, sticky="w", padx=7, pady=7)
        txtColumnLetter.grid(row=2, column=1, padx=10, pady=10)
        lblFirstRow.grid(row=3, column=0, sticky="w", padx=7, pady=7)
        txtFirstRow.grid(row=3, column=1, padx=7, pady=7)
        lblLastRow.grid(row=4, column=0, sticky="w", padx=7, pady=7)
        txtLastRow.grid(row=4, column=1, padx=7, pady=7)
        browse_button.grid(row=0, column=2, padx=7, pady=7)
        submit_button.grid(row=4, column=2, padx=7, pady=7)

        root.resizable(False, False)
        root.mainloop()

        # endregion UI Prompts
        self.file_name = Automator.create_excel(self)

    def run_test1(self):
        try:
            # Connect the application to Label Maintenance and send confirmation message
            app = Application(backend="uia").connect(title="Label Maintenance")
            print('Connection to Label Maintenance achieved')

            # Load in a specific workbook
            book = load_workbook(user_data[0])

            # Load in user data
            last_row = int(user_data[4])
            first_row = int(user_data[3])
            col_letter = user_data[2]
            n = user_data[1]
            sheets = book.sheetnames
            ws = book[sheets[n]]

            # Loop through the given parts and read/write the part number into Label Maintenance
            # The first number in the range function is the starting row
            # The second number is the final row
            for i in range(first_row, last_row + 1):

                app = Application(backend="uia").connect(title="Label Maintenance")
                main_window = app.window(title='Label Maintenance')

                # Find the cell and get its value
                cell = col_letter + str(i)
                cell_value = ws[cell].value

                # Type cell value into text box
                main_window.child_window(auto_id='txtKeyField').type_keys(cell_value)
                send_keys("{TAB}")

                # Check for invalid part numbers
                if main_window.child_window(title='Error').exists(timeout=10):
                    main_window.child_window(title='OK').click()
                    if main_window.child_window(title='Error').exists(timeout=5):
                        main_window.child_window(title='OK').click()
                    # Write invalid part into Excel file
                    Automator.write_invalid(self, cell_value, self.file_name)
                    print(cell_value + " - Invalid Part Number")
                    continue

                # Confirms label does not exist
                if main_window.child_window(title="Add New Confirmation").exists(timeout=5):

                    main_window.child_window(auto_id='btnYes2').click_input()

                    # Check for invalid part numbers
                    if main_window.child_window(title='Error').exists(timeout=10):
                        main_window.child_window(title='OK').click()
                        if main_window.child_window(title='Error').exists(timeout=5):
                            main_window.child_window(title='OK').click()

                        # Write invalid part number in Excel sheet
                        Automator.write_invalid(self, cell_value, self.file_name)
                        print(str(cell_value) + " - Invalid Part Number")
                        continue

                    else:
                        # Save if part number is valid
                        main_window.child_window(title="Save").click_input()
                        main_window.child_window(title="Clear").click_input()
                        print(str(cell_value) + " - Label Created")
                else:
                    Automator.write_exists(self, cell_value, self.file_name)
                    main_window.child_window(title="Clear").click_input()
                    print(str(cell_value) + " - Label Already Exists")

        except pywinauto.findwindows.ElementNotFoundError:
            try:
                self.run_test1()
            except pywinauto.findwindows.ElementNotFoundError:
                print("Connection Failed. Please be sure to have Label Maintenance running")
                response = input("Try again? (yes or no)")
                if response == "yes":
                    self.run_test1()
        except pywinauto.timings.TimeoutError:
            print("The program took too long to respond. Please restart")
        except Exception as e:
            print(e)
            raise e

        print("Session summary is saved in Epicor Automation -> build")

    def reconnect(self, app_name, timeout=30):
        try:
            # Try connecting to the application window
            Timings.app_start_timeout = timeout
            app = pywinauto.Application().connect(path=None, title=app_name)
            return True
        except pywinauto.findwindows.ElementNotFoundError:
            return False

    # region File Reading
    def create_file_name(self) -> str:
        # Save the workbook with a unique name (e.g., based on timestamp)
        file_name = f"output_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return file_name

    def create_excel(self) -> str:
        # Write data to the pre-created sheet (global variable)
        return_sheet['A1'] = 'Invalid PN'
        return_sheet['B1'] = 'Already Had Label'

        # Give the sheet a unique name and save it
        file_name = Automator.create_file_name(self)
        return_workbook.save(file_name)
        print(f"Excel file '{file_name}' created successfully.")

        return file_name

    def write_invalid(self, part_number, file_name):
        # Write PN into Excel file
        return_sheet['A' + str(self.invalid_index)] = part_number
        self.invalid_index += 1
        return_workbook.save(file_name)

    def write_exists(self, part_number, file_name):
        # Write PN into Excel file
        return_sheet['B' + str(self.existed_index)] = part_number
        self.existed_index += 1
        return_workbook.save(file_name)

    # endregion File Reading
